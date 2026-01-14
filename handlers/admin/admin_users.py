"""
Admin Users Handler
===================
Foydalanuvchilarni boshqarish handlerlari
"""

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from datetime import datetime, timedelta

from loader import dp, bot, user_db
from keyboards.inline.admin_keyboards import users_menu, back_button
from keyboards.default.admin_keyboards import admin_cancel_button, admin_skip_button, remove_keyboard
from states.admin_states import UserManageStates,AccessManageStates
from handlers.admin.admin_start import admin_required


# ============================================================
#                    FOYDALANUVCHILAR MENYUSI
# ============================================================

@dp.callback_query_handler(text="admin:users")
@admin_required
async def show_users_menu(call: types.CallbackQuery):
    """Foydalanuvchilar menyusi"""

    # Statistika
    total = user_db.execute("SELECT COUNT(*) FROM Users", fetchone=True)

    today = datetime.now().strftime('%Y-%m-%d')
    today_new = user_db.execute(
        "SELECT COUNT(*) FROM Users WHERE DATE(created_at) = ?",
        parameters=(today,),
        fetchone=True
    )

    paid = user_db.execute(
        """SELECT COUNT(DISTINCT user_id) FROM Payments WHERE status = 'approved'""",
        fetchone=True
    )

    text = f"""
👥 <b>Foydalanuvchilar boshqaruvi</b>

📊 <b>Statistika:</b>
├ 👥 Jami: <b>{total[0] if total else 0}</b>
├ 🆕 Bugun: <b>{today_new[0] if today_new else 0}</b>
└ 💰 Pullik: <b>{paid[0] if paid else 0}</b>

⬇️ Bo'limni tanlang:
"""

    await call.message.edit_text(text, reply_markup=users_menu())
    await call.answer()


# ============================================================
#                    BARCHA FOYDALANUVCHILAR
# ============================================================

@dp.callback_query_handler(text="admin:users:all")
@admin_required
async def show_all_users(call: types.CallbackQuery):
    """Barcha foydalanuvchilar"""

    users = user_db.execute(
        """SELECT id, telegram_id, full_name, phone, created_at
           FROM Users
           ORDER BY created_at DESC
           LIMIT 20""",
        fetchall=True
    )

    if not users:
        await call.answer("📭 Foydalanuvchilar yo'q", show_alert=True)
        return

    text = f"""
👥 <b>Barcha foydalanuvchilar</b>

So'nggi 20 ta:

"""

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    for u in users:
        name = u[2] or "Noma'lum"
        date = u[4][:10] if u[4] else ""

        keyboard.add(types.InlineKeyboardButton(
            f"👤 {name}",
            callback_data=f"admin:user:view:{u[0]}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "🔍 Qidirish",
        callback_data="admin:users:search"
    ))
    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users"
    ))

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


# ============================================================
#                    FOYDALANUVCHI QIDIRISH
# ============================================================

@dp.callback_query_handler(text="admin:users:search")
@admin_required
async def search_user_start(call: types.CallbackQuery, state: FSMContext):
    """Foydalanuvchi qidirish"""

    await call.message.edit_text(
        "🔍 <b>Foydalanuvchi qidirish</b>\n\n"
        "Qidirish uchun quyidagilardan birini kiriting:\n"
        "• Telegram ID\n"
        "• Telefon raqam\n"
        "• Ism\n"
        "• Username"
    )

    await call.message.answer(
        "✍️ Qidirish so'zini kiriting:",
        reply_markup=admin_cancel_button()
    )

    await UserManageStates.search.set()
    await call.answer()


@dp.message_handler(state=UserManageStates.search)
async def search_user_process(message: types.Message, state: FSMContext):
    """Qidirish natijasi"""
    if message.text == "❌ Bekor qilish":
        await state.finish()
        await message.answer("❌ Bekor qilindi", reply_markup=remove_keyboard())
        return

    query = message.text.strip()

    # Qidirish
    users = user_db.execute(
        """SELECT id, telegram_id, full_name, phone, username
           FROM Users
           WHERE telegram_id LIKE ? 
              OR phone LIKE ? 
              OR full_name LIKE ? 
              OR username LIKE ?
           LIMIT 10""",
        parameters=(f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"),
        fetchall=True
    )

    await state.finish()

    if not users:
        await message.answer(
            f"📭 <b>\"{query}\"</b> bo'yicha hech narsa topilmadi.",
            reply_markup=remove_keyboard()
        )
        return

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    text = f"🔍 <b>Qidirish natijasi:</b> \"{query}\"\n\n"
    text += f"Topildi: {len(users)} ta\n\n"

    for u in users:
        name = u[2] or "Noma'lum"
        keyboard.add(types.InlineKeyboardButton(
            f"👤 {name} | {u[1]}",
            callback_data=f"admin:user:view:{u[0]}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "🔍 Qayta qidirish",
        callback_data="admin:users:search"
    ))
    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users"
    ))

    await message.answer(text, reply_markup=keyboard)


@dp.message_handler(commands=['delete_all_users'])
@admin_required
async def delete_all_users_command(message: types.Message):
    """
    Barcha foydalanuvchilarni o'chirish
    Format: /delete_all_users TASDIQLASH
    """
    args = message.text.split()

    if len(args) != 2 or args[1] != "TASDIQLASH":
        count = user_db.count_users()
        await message.answer(
            f"⚠️ <b>Diqqat!</b>\n\n"
            f"Bu buyruq {count} ta foydalanuvchini o'chiradi!\n"
            f"(Adminlar o'chirilmaydi)\n\n"
            f"Tasdiqlash uchun:\n"
            f"<code>/delete_all_users TASDIQLASH</code>"
        )
        return

    count = user_db.delete_all_users()

    await message.answer(f"✅ {count} ta foydalanuvchi o'chirildi!")





# ============================================================
#                    FOYDALANUVCHINI O'CHIRISH (COMMAND)
# ============================================================

@dp.message_handler(commands=['delete_user'])
@admin_required
async def delete_user_command(message: types.Message):
    """
    Foydalanuvchini o'chirish
    Format: /delete_user 123456789
    """
    args = message.text.split()

    if len(args) != 2:
        await message.answer(
            "❌ Noto'g'ri format!\n\n"
            "✅ To'g'ri: /delete_user 123456789\n"
            "(telegram_id kiriting)"
        )
        return

    try:
        telegram_id = int(args[1])
    except ValueError:
        await message.answer("❌ Telegram ID faqat raqam bo'lishi kerak!")
        return

    # Foydalanuvchini tekshirish
    user = user_db.get_user(telegram_id)

    if not user:
        await message.answer(f"❌ Foydalanuvchi topilmadi: {telegram_id}")
        return

    # O'chirish
    user_db.delete_user(telegram_id)

    await message.answer(
        f"✅ Foydalanuvchi o'chirildi!\n\n"
        f"👤 {user['full_name'] or 'Nomalum'}\n"
        f"🆔 {telegram_id}"
    )





# ============================================================
#                    FOYDALANUVCHINI KO'RISH
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:view:")
@admin_required
async def view_user(call: types.CallbackQuery):
    """Foydalanuvchini ko'rish"""
    user_id = int(call.data.split(":")[-1])

    user = user_db.execute(
        """SELECT * FROM Users WHERE id = ?""",
        parameters=(user_id,),
        fetchone=True
    )

    if not user:
        await call.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)
        return

    # user tuple:
    # 0-id, 1-telegram_id, 2-username, 3-full_name, 4-phone
    # 5-total_score, 6-created_at, 7-last_activity

    # Kurslar
    courses = user_db.execute(
        """SELECT COUNT(DISTINCT course_id) FROM Payments 
           WHERE user_id = ? AND status = 'approved'""",
        parameters=(user_id,),
        fetchone=True
    )

    # To'lovlar
    payments = user_db.execute(
        """SELECT SUM(amount) FROM Payments 
           WHERE user_id = ? AND status = 'approved'""",
        parameters=(user_id,),
        fetchone=True
    )

    # Tugatilgan darslar
    completed = user_db.execute(
        """SELECT COUNT(*) FROM UserProgress 
           WHERE user_id = ? AND status = 'completed'""",
        parameters=(user_id,),
        fetchone=True
    )

    text = f"""
👤 <b>Foydalanuvchi</b>

🆔 ID: <code>{user[1]}</code>
👤 Ism: <b>{user[3] or 'Nomalum'}</b>
📱 Telefon: {user[4] or 'Yoq'}
💬 Username: @{user[2] or 'yoq'}

📊 <b>Statistika:</b>
├ ⭐️ Ball: {user[5] or 0}
├ 📚 Kurslar: {courses[0] if courses else 0}
├ 💰 To'lovlar: {payments[0]:,.0f if payments and payments[0] else 0} so'm
└ ✅ Darslar: {completed[0] if completed else 0}

📅 Ro'yxatdan o'tgan: {user[6][:10] if user[6] else 'Nomalum'}
🕐 Oxirgi faollik: {user[7][:16] if user[7] else 'Nomalum'}
"""

    keyboard = types.InlineKeyboardMarkup(row_width=2)

    keyboard.add(
        types.InlineKeyboardButton(
            "📚 Kurslar",
            callback_data=f"admin:user:courses:{user_id}"
        ),
        types.InlineKeyboardButton(
            "💰 To'lovlar",
            callback_data=f"admin:user:payments:{user_id}"
        )
    )
    keyboard.add(
        types.InlineKeyboardButton(
            "🔓 Dostup berish",
            callback_data=f"admin:user:access:{user_id}"
        ),
        types.InlineKeyboardButton(
            "💬 Xabar yuborish",
            callback_data=f"admin:user:message:{user_id}"
        )
    )
    keyboard.add(
        types.InlineKeyboardButton(
            "⭐️ Ball qo'shish",
            callback_data=f"admin:user:score:{user_id}"
        ),
        types.InlineKeyboardButton(
            "🚫 Bloklash",
            callback_data=f"admin:user:block:{user_id}"
        )
    )
    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users:all"
    ))

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


# ============================================================
#                    DOSTUP BERISH
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:access:")
@admin_required
async def give_access_start(call: types.CallbackQuery, state: FSMContext):
    """Dostup berish - kurs tanlash"""
    user_id = int(call.data.split(":")[-1])

    courses = user_db.get_all_courses(active_only=True)

    if not courses:
        await call.answer("📭 Kurslar yo'q", show_alert=True)
        return

    await state.update_data(target_user_id=user_id)

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    for course in courses:
        keyboard.add(types.InlineKeyboardButton(
            f"📚 {course['name']}",
            callback_data=f"admin:access:course:{course['id']}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data=f"admin:user:view:{user_id}"
    ))

    await call.message.edit_text(
        "🔓 <b>Dostup berish</b>\n\n"
        "Kursni tanlang:",
        reply_markup=keyboard
    )
    await call.answer()


@dp.callback_query_handler(text_startswith="admin:access:course:")
@admin_required
async def give_access_course(call: types.CallbackQuery, state: FSMContext):
    """Dostup berish - tasdiqlash"""
    course_id = int(call.data.split(":")[-1])

    data = await state.get_data()
    user_id = data.get('target_user_id')

    if not user_id:
        await call.answer("❌ Xatolik!", show_alert=True)
        return

    course = user_db.get_course(course_id)

    # Dostup berish
    user_db.give_manual_access(
        user_id=user_id,
        course_id=course_id,
        admin_id=user_db.get_admin_id(call.from_user.id),
        reason="Admin tomonidan berildi"
    )

    await state.finish()

    # Foydalanuvchiga xabar
    user = user_db.execute(
        "SELECT telegram_id FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    if user:
        try:
            await bot.send_message(
                user[0],
                f"🎉 Sizga <b>{course['name']}</b> kursiga dostup berildi!\n\n"
                f"\"📚 Mening kurslarim\" bo'limiga kiring."
            )
        except:
            pass

    await call.message.edit_text(
        f"✅ <b>Dostup berildi!</b>\n\n"
        f"📚 Kurs: {course['name']}\n"
        f"👤 Foydalanuvchi ID: {user_id}",
        reply_markup=back_button(f"admin:user:view:{user_id}")
    )
    await call.answer()


# ============================================================
#                    XABAR YUBORISH
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:message:")
@admin_required
async def send_message_start(call: types.CallbackQuery, state: FSMContext):
    """Xabar yuborish"""
    user_id = int(call.data.split(":")[-1])

    await state.update_data(target_user_id=user_id)

    await call.message.edit_text(
        "💬 <b>Xabar yuborish</b>\n\n"
        "Xabar matnini kiriting:"
    )

    await call.message.answer(
        "✍️ Xabar yozing:",
        reply_markup=admin_cancel_button()
    )

    await UserManageStates.send_message.set()
    await call.answer()


@dp.message_handler(state=UserManageStates.send_message)
async def send_message_process(message: types.Message, state: FSMContext):
    """Xabarni yuborish"""
    if message.text == "❌ Bekor qilish":
        await state.finish()
        await message.answer("❌ Bekor qilindi", reply_markup=remove_keyboard())
        return

    data = await state.get_data()
    user_id = data.get('target_user_id')

    user = user_db.execute(
        "SELECT telegram_id FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    await state.finish()

    if not user:
        await message.answer("❌ Foydalanuvchi topilmadi!", reply_markup=remove_keyboard())
        return

    try:
        await bot.send_message(
            user[0],
            f"📩 <b>Admin xabari:</b>\n\n{message.text}"
        )
        await message.answer("✅ Xabar yuborildi!", reply_markup=remove_keyboard())
    except Exception as e:
        await message.answer(f"❌ Xabar yuborib bo'lmadi: {e}", reply_markup=remove_keyboard())


# ============================================================
#                    BALL QO'SHISH
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:score:")
@admin_required
async def add_score_start(call: types.CallbackQuery, state: FSMContext):
    """Ball qo'shish"""
    user_id = int(call.data.split(":")[-1])

    await state.update_data(target_user_id=user_id)

    await call.message.edit_text(
        "⭐️ <b>Ball qo'shish</b>\n\n"
        "Qo'shiladigan ballni kiriting (musbat yoki manfiy son):"
    )

    await call.message.answer(
        "✍️ Ball kiriting (masalan: 50 yoki -20):",
        reply_markup=admin_cancel_button()
    )

    await UserManageStates.add_score.set()
    await call.answer()


@dp.message_handler(state=UserManageStates.add_score)
async def add_score_process(message: types.Message, state: FSMContext):
    """Ballni qo'shish"""
    if message.text == "❌ Bekor qilish":
        await state.finish()
        await message.answer("❌ Bekor qilindi", reply_markup=remove_keyboard())
        return

    try:
        score = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Noto'g'ri format! Son kiriting.")
        return

    data = await state.get_data()
    user_id = data.get('target_user_id')

    user = user_db.execute(
        "SELECT telegram_id, total_score FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    if not user:
        await state.finish()
        await message.answer("❌ Foydalanuvchi topilmadi!", reply_markup=remove_keyboard())
        return

    new_score = (user[1] or 0) + score

    user_db.execute(
        "UPDATE Users SET total_score = ? WHERE id = ?",
        parameters=(new_score, user_id)
    )

    await state.finish()

    sign = "+" if score > 0 else ""
    await message.answer(
        f"✅ Ball qo'shildi!\n\n"
        f"Qo'shilgan: {sign}{score}\n"
        f"Yangi ball: {new_score}",
        reply_markup=remove_keyboard()
    )


# ============================================================
#                    PULLIK FOYDALANUVCHILAR
# ============================================================

@dp.callback_query_handler(text="admin:users:paid")
@admin_required
async def show_paid_users(call: types.CallbackQuery):
    """Pullik foydalanuvchilar"""

    users = user_db.execute(
        """SELECT u.id, u.full_name, u.phone, 
                  COUNT(DISTINCT p.course_id) as courses,
                  SUM(p.amount) as total_paid
           FROM Users u
           JOIN Payments p ON u.id = p.user_id AND p.status = 'approved'
           GROUP BY u.id
           ORDER BY total_paid DESC
           LIMIT 20""",
        fetchall=True
    )

    if not users:
        await call.answer("📭 Pullik foydalanuvchilar yo'q", show_alert=True)
        return

    text = f"""
💰 <b>Pullik foydalanuvchilar</b>

Top 20:

"""

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    for i, u in enumerate(users, 1):
        name = u[1] or "Noma'lum"
        text += f"{i}. {name} | {u[3]} kurs | {u[4]:,.0f} so'm\n"

        keyboard.add(types.InlineKeyboardButton(
            f"👤 {name}",
            callback_data=f"admin:user:view:{u[0]}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users"
    ))

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


# ============================================================
#                    TOP FOYDALANUVCHILAR
# ============================================================

@dp.callback_query_handler(text="admin:users:top")
@admin_required
async def show_top_users(call: types.CallbackQuery):
    """Top foydalanuvchilar (ball bo'yicha)"""

    users = user_db.execute(
        """SELECT id, full_name, total_score
           FROM Users
           WHERE total_score > 0
           ORDER BY total_score DESC
           LIMIT 20""",
        fetchall=True
    )

    if not users:
        await call.answer("📭 Natijalar yo'q", show_alert=True)
        return

    text = f"""
🏆 <b>Top foydalanuvchilar</b>

Ball bo'yicha Top 20:

"""

    medals = ['🥇', '🥈', '🥉']

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    for i, u in enumerate(users, 1):
        name = u[1] or "Noma'lum"
        medal = medals[i - 1] if i <= 3 else f"{i}."
        text += f"{medal} {name} - {u[2]} ball\n"

        keyboard.add(types.InlineKeyboardButton(
            f"👤 {name}",
            callback_data=f"admin:user:view:{u[0]}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users"
    ))

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


# ============================================================
#                    FOYDALANUVCHI KURSLARI
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:courses:")
@admin_required
async def show_user_courses(call: types.CallbackQuery):
    """Foydalanuvchi kurslari"""
    user_id = int(call.data.split(":")[-1])

    courses = user_db.execute(
        """SELECT c.name, p.amount, p.status, p.approved_at
           FROM Payments p
           JOIN Courses c ON p.course_id = c.id
           WHERE p.user_id = ?
           ORDER BY p.created_at DESC""",
        parameters=(user_id,),
        fetchall=True
    )

    # Manual access
    manual = user_db.execute(
        """SELECT c.name, ma.created_at, ma.expires_at, ma.reason
           FROM ManualAccess ma
           JOIN Courses c ON ma.course_id = c.id
           WHERE ma.user_id = ?""",
        parameters=(user_id,),
        fetchall=True
    )

    text = f"📚 <b>Foydalanuvchi kurslari</b>\n\n"

    if courses:
        text += "<b>To'lov orqali:</b>\n"
        for c in courses:
            status = "✅" if c[2] == 'approved' else "⏳" if c[2] == 'pending' else "❌"
            text += f"{status} {c[0]} | {c[1]:,.0f} so'm\n"

    if manual:
        text += "\n<b>Admin tomonidan:</b>\n"
        for m in manual:
            text += f"🔓 {m[0]}"
            if m[3]:
                text += f" ({m[3]})"
            text += "\n"

    if not courses and not manual:
        text += "📭 Kurslar yo'q"

    await call.message.edit_text(
        text,
        reply_markup=back_button(f"admin:user:view:{user_id}")
    )
    await call.answer()


# ============================================================
#                    FOYDALANUVCHI TO'LOVLARI
# ============================================================

@dp.callback_query_handler(text_startswith="admin:user:payments:")
@admin_required
async def show_user_payments(call: types.CallbackQuery):
    """Foydalanuvchi to'lovlari"""
    user_id = int(call.data.split(":")[-1])

    payments = user_db.execute(
        """SELECT p.id, p.amount, p.status, p.created_at, c.name
           FROM Payments p
           JOIN Courses c ON p.course_id = c.id
           WHERE p.user_id = ?
           ORDER BY p.created_at DESC
           LIMIT 10""",
        parameters=(user_id,),
        fetchall=True
    )

    if not payments:
        await call.answer("📭 To'lovlar yo'q", show_alert=True)
        return

    status_map = {'pending': '⏳', 'approved': '✅', 'rejected': '❌'}

    text = f"💰 <b>Foydalanuvchi to'lovlari</b>\n\n"

    for p in payments:
        status = status_map.get(p[2], '❓')
        date = p[3][:10] if p[3] else ""
        text += f"{status} #{p[0]} | {p[4]}\n   {p[1]:,.0f} so'm | {date}\n\n"

    await call.message.edit_text(
        text,
        reply_markup=back_button(f"admin:user:view:{user_id}")
    )
    await call.answer()


# ============================================================
#                    SOTIB OLMAGANLAR (PAGINATION BILAN)
# ============================================================

@dp.callback_query_handler(text="admin:users:not_paid")
@admin_required
async def show_not_paid_users(call: types.CallbackQuery):
    """Sotib olmaganlar - 1-sahifa"""
    await show_not_paid_page(call, page=0)


@dp.callback_query_handler(text_startswith="admin:users:not_paid:")
@admin_required
async def show_not_paid_users_page(call: types.CallbackQuery):
    """Sotib olmaganlar - pagination"""
    page = int(call.data.split(":")[-1])
    await show_not_paid_page(call, page)


async def show_not_paid_page(call: types.CallbackQuery, page: int = 0):
    """Sotib olmaganlar ro'yxati (pagination bilan)"""

    per_page = 10  # Har sahifada 10 ta
    offset = page * per_page

    # Jami sonini olish
    total_result = user_db.execute(
        """SELECT COUNT(*) FROM Users u
           WHERE u.phone IS NOT NULL
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM Payments WHERE status = 'approved'
           )
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM ManualAccess WHERE is_active = 1
           )""",
        fetchone=True
    )
    total = total_result[0] if total_result else 0

    # Userlarni olish
    users = user_db.execute(
        """SELECT u.id, u.telegram_id, u.full_name, u.phone, u.username, u.created_at
           FROM Users u
           WHERE u.phone IS NOT NULL
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM Payments WHERE status = 'approved'
           )
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM ManualAccess WHERE is_active = 1
           )
           ORDER BY u.created_at DESC
           LIMIT ? OFFSET ?""",
        parameters=(per_page, offset),
        fetchall=True
    )

    if not users and page == 0:
        await call.answer("📭 Bunday foydalanuvchilar yo'q", show_alert=True)
        return

    total_pages = (total + per_page - 1) // per_page  # Jami sahifalar

    text = f"""
🆕 <b>Sotib olmaganlar</b>

Ro'yxatdan o'tgan, lekin kurs olmagan:
📊 Jami: <b>{total}</b> ta
📄 Sahifa: <b>{page + 1}/{total_pages}</b>

"""

    # Har bir user ma'lumotlari
    for i, u in enumerate(users, start=offset + 1):
        user_id = u[0]
        telegram_id = u[1]
        name = u[2] or "Noma'lum"
        phone = u[3] or "Yo'q"
        username = f"@{u[4]}" if u[4] else "yo'q"
        date = u[5][:10] if u[5] else ""

        text += f"""
<b>{i}. {name}</b>
   📱 {phone}
   💬 {username}
   📅 {date}
"""

    keyboard = types.InlineKeyboardMarkup(row_width=2)

    # Pagination tugmalari
    nav_buttons = []

    if page > 0:
        nav_buttons.append(
            types.InlineKeyboardButton("⬅️ Oldingi", callback_data=f"admin:users:not_paid:{page - 1}")
        )

    if page < total_pages - 1:
        nav_buttons.append(
            types.InlineKeyboardButton("Keyingi ➡️", callback_data=f"admin:users:not_paid:{page + 1}")
        )

    if nav_buttons:
        keyboard.row(*nav_buttons)

    # Export tugmasi (katta ro'yxat uchun)
    if total > 10:
        keyboard.add(
            types.InlineKeyboardButton("📥 Excel yuklash", callback_data="admin:users:not_paid:export")
        )

    keyboard.add(
        types.InlineKeyboardButton("⬅️ Orqaga", callback_data="admin:users")
    )

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


# ============================================================
#                    EXCEL EXPORT
# ============================================================

@dp.callback_query_handler(text="admin:users:not_paid:export")
@admin_required
async def export_not_paid_users(call: types.CallbackQuery):
    """Sotib olmaganlarni Excel ga export qilish"""

    await call.answer("⏳ Fayl tayyorlanmoqda...", show_alert=False)

    users = user_db.execute(
        """SELECT u.full_name, u.phone, u.username, u.telegram_id, u.created_at
           FROM Users u
           WHERE u.phone IS NOT NULL
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM Payments WHERE status = 'approved'
           )
           AND u.id NOT IN (
               SELECT DISTINCT user_id FROM ManualAccess WHERE is_active = 1
           )
           ORDER BY u.created_at DESC""",
        fetchall=True
    )

    if not users:
        await call.answer("📭 Ma'lumot yo'q", show_alert=True)
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sotib olmaganlar"

        # Sarlavhalar
        headers = ["№", "Ism Familya", "Telefon", "Username", "Telegram ID", "Ro'yxatdan o'tgan"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Ma'lumotlar
        for row, u in enumerate(users, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=u[0] or "Noma'lum")
            ws.cell(row=row, column=3, value=u[1] or "")
            ws.cell(row=row, column=4, value=f"@{u[2]}" if u[2] else "")
            ws.cell(row=row, column=5, value=u[3])
            ws.cell(row=row, column=6, value=u[4][:10] if u[4] else "")

        # Ustun kengligini sozlash
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Faylni saqlash
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        from aiogram.types import InputFile
        from datetime import datetime

        filename = f"sotib_olmaganlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        await call.message.answer_document(
            InputFile(file_stream, filename=filename),
            caption=f"📥 <b>Sotib olmaganlar ro'yxati</b>\n\nJami: {len(users)} ta"
        )

    except ImportError:
        await call.message.answer("❌ Excel yaratish uchun openpyxl kerak!")
    except Exception as e:
        await call.message.answer(f"❌ Xatolik: {e}")


# ============================================================
#                    DOSTUP BOSHQARUVI (YANGI)
# ============================================================

from states.admin_states import AccessManageStates


@dp.callback_query_handler(text="admin:users:access_manage")
@admin_required
async def access_manage_start(call: types.CallbackQuery):
    """Dostup boshqaruvi - bosh sahifa"""

    text = """
🔐 <b>Dostup boshqaruvi</b>

Bu yerda istalgan foydalanuvchiga:
- 🔓 Kursga dostup ochish
- 🔒 Dostupni yopish

Foydalanuvchini qidiring:
- Telegram ID
- Telefon raqam
- Ism
- Username
"""

    keyboard = types.InlineKeyboardMarkup(row_width=1)
    keyboard.add(
        types.InlineKeyboardButton("🔍 Foydalanuvchi qidirish", callback_data="admin:access:search")
    )
    keyboard.add(
        types.InlineKeyboardButton("⬅️ Orqaga", callback_data="admin:users")
    )

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


@dp.callback_query_handler(text="admin:access:search")
@admin_required
async def access_search_start(call: types.CallbackQuery, state: FSMContext):
    """Dostup uchun user qidirish"""

    await call.message.edit_text(
        "🔍 <b>Foydalanuvchi qidirish</b>\n\n"
        "Qidirish uchun kiriting:\n"
        "• Telegram ID\n"
        "• Telefon raqam\n"
        "• Ism\n"
        "• Username"
    )

    await call.message.answer(
        "✍️ Qidirish so'zini kiriting:",
        reply_markup=admin_cancel_button()
    )

    await AccessManageStates.search.set()
    await call.answer()


@dp.message_handler(state=AccessManageStates.search)
async def access_search_process(message: types.Message, state: FSMContext):
    """Qidirish natijasi - dostup uchun"""
    if message.text == "❌ Bekor qilish":
        await state.finish()
        await message.answer("❌ Bekor qilindi", reply_markup=remove_keyboard())
        return

    query = message.text.strip()

    # Qidirish
    users = user_db.execute(
        """SELECT id, telegram_id, full_name, phone, username
           FROM Users
           WHERE telegram_id LIKE ? 
              OR phone LIKE ? 
              OR full_name LIKE ? 
              OR username LIKE ?
           LIMIT 10""",
        parameters=(f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"),
        fetchall=True
    )

    await state.finish()

    if not users:
        await message.answer(
            f"📭 <b>\"{query}\"</b> bo'yicha hech narsa topilmadi.",
            reply_markup=remove_keyboard()
        )
        return

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    text = f"🔍 <b>Topildi:</b> {len(users)} ta\n\n"

    for u in users:
        user_id = u[0]
        name = u[2] or "Noma'lum"
        phone = u[3] or ""

        btn_text = f"👤 {name}"
        if phone:
            btn_text += f" | {phone}"

        keyboard.add(types.InlineKeyboardButton(
            btn_text,
            callback_data=f"admin:access:user:{user_id}"
        ))

    keyboard.add(types.InlineKeyboardButton(
        "🔍 Qayta qidirish",
        callback_data="admin:access:search"
    ))
    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users:access_manage"
    ))

    await message.answer(text, reply_markup=keyboard)


@dp.callback_query_handler(text_startswith="admin:access:user:")
@admin_required
async def access_user_courses(call: types.CallbackQuery):
    """User kurslarini ko'rsatish - dostup bilan"""
    user_id = int(call.data.split(":")[-1])

    # User ma'lumotlari
    user = user_db.execute(
        "SELECT telegram_id, full_name, phone FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    if not user:
        await call.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)
        return

    telegram_id = user[0]
    user_name = user[1] or "Noma'lum"
    phone = user[2] or "Yo'q"

    # Barcha kurslar
    courses = user_db.get_all_courses(active_only=True)

    text = f"""
🔐 <b>Dostup boshqaruvi</b>

👤 <b>{user_name}</b>
📱 {phone}
🆔 <code>{telegram_id}</code>

📚 <b>Kurslar:</b>
"""

    keyboard = types.InlineKeyboardMarkup(row_width=1)

    for course in courses:
        course_id = course['id']
        course_name = course['name']

        # Dostup bormi tekshirish
        has_access = user_db.has_course_access(telegram_id, course_id)

        if has_access:
            status = "✅"
            btn_text = f"{status} {course_name} | 🔒 Yopish"
            callback = f"admin:access:close:{user_id}:{course_id}"
        else:
            status = "❌"
            btn_text = f"{status} {course_name} | 🔓 Ochish"
            callback = f"admin:access:open:{user_id}:{course_id}"

        keyboard.add(types.InlineKeyboardButton(btn_text, callback_data=callback))

    keyboard.add(types.InlineKeyboardButton(
        "🔍 Boshqa user",
        callback_data="admin:access:search"
    ))
    keyboard.add(types.InlineKeyboardButton(
        "⬅️ Orqaga",
        callback_data="admin:users:access_manage"
    ))

    await call.message.edit_text(text, reply_markup=keyboard)
    await call.answer()


@dp.callback_query_handler(text_startswith="admin:access:open:")
@admin_required
async def access_open_course(call: types.CallbackQuery):
    """Kursga dostup ochish"""
    parts = call.data.split(":")
    user_id = int(parts[3])
    course_id = int(parts[4])

    # User telegram_id
    user = user_db.execute(
        "SELECT telegram_id, full_name FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    if not user:
        await call.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)
        return

    telegram_id = user[0]
    user_name = user[1] or "Noma'lum"

    # Kurs
    course = user_db.get_course(course_id)
    if not course:
        await call.answer("❌ Kurs topilmadi!", show_alert=True)
        return

    # Dostup berish
    success = user_db.grant_manual_access(
        telegram_id=telegram_id,
        course_id=course_id,
        admin_telegram_id=call.from_user.id,
        reason="Admin tomonidan berildi"
    )

    if success:
        # Userga xabar
        try:
            await bot.send_message(
                telegram_id,
                f"🎉 Sizga <b>{course['name']}</b> kursiga dostup berildi!\n\n"
                f"📚 \"Mening kurslarim\" bo'limiga kiring."
            )
        except:
            pass

        await call.answer(f"✅ {user_name} ga dostup berildi!", show_alert=True)
    else:
        await call.answer("❌ Xatolik!", show_alert=True)

    # Qayta yuklash - call.data ni o'zgartiramiz
    call.data = f"admin:access:user:{user_id}"
    await access_user_courses(call)


@dp.callback_query_handler(text_startswith="admin:access:close:")
@admin_required
async def access_close_course(call: types.CallbackQuery):
    """Kurs dostupini yopish"""
    parts = call.data.split(":")
    user_id = int(parts[3])
    course_id = int(parts[4])

    # User telegram_id
    user = user_db.execute(
        "SELECT telegram_id, full_name FROM Users WHERE id = ?",
        parameters=(user_id,),
        fetchone=True
    )

    if not user:
        await call.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)
        return

    telegram_id = user[0]
    user_name = user[1] or "Noma'lum"

    # Kurs
    course = user_db.get_course(course_id)
    if not course:
        await call.answer("❌ Kurs topilmadi!", show_alert=True)
        return

    # Dostupni yopish
    user_db.revoke_access(telegram_id, course_id)

    # Userga xabar
    try:
        await bot.send_message(
            telegram_id,
            f"⚠️ Sizning <b>{course['name']}</b> kursiga dostupingiz yopildi.\n\n"
            f"Savollar bo'lsa admin bilan bog'laning."
        )
    except:
        pass

    await call.answer(f"🔒 {user_name} dostup yopildi!", show_alert=True)

    # Qayta yuklash
    call.data = f"admin:access:user:{user_id}"
    await access_user_courses(call)